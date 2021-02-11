import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine
import datetime
import time

class ReadandWrite:
    def __init__(self):
        self.engine = create_engine('mysql+pymysql://root:root@localhost:3306/vidrio_test')
        self.con = self.engine.connect()

    def excel_to_db(self, filenames):
        for filename in filenames:
            wb = load_workbook(filename)
            xls = pd.ExcelFile(filename)
            dataframes = []
            for worksheet in wb.sheetnames:
                dataframes.append((pd.read_excel(xls, worksheet), worksheet))

            for df, name in dataframes:
                df.columns = df.columns.str.strip()
                name = name.strip().replace(' ', '_')
                df.to_sql(name, con=self.con)
                # print(df)

    def generate_portfolio_valuation_file1(self, old_filenames, new_filename):
        input_reference_date = self.extract_timestamp_from_file(old_filenames[0])

        dataframes = []
        for filename in old_filenames:
            wb = load_workbook(filename)
            xls = pd.ExcelFile(filename)

            for worksheet in wb.sheetnames:
                dataframes.append(pd.read_excel(xls, worksheet))

        for df in dataframes:  
            df.columns = df.columns.str.strip()
        
        data = []
        #dataframes[0] = constituents
        #dataframes[1] = index data
        #dataframes[2] = mapping data
        for index, row in dataframes[0].iterrows():
            new_row = []
            if input_reference_date.replace('/', '-') in str(row['Date']):
                reference_day = row['Date']
                periodicty = 'Daily'
                investor_account_uid = dataframes[1]['ISIN'].loc[lambda x : x == 'HFRIILAU'][0]
                investor_account_long_name = dataframes[1].loc[dataframes[1].index[dataframes[1]['ISIN'].loc[lambda x : x=='HFRIILAU'].index[0]], 'Index Name']
                investment_account_uid = None
                matching_indices = dataframes[2]['Counterparty ID'].loc[lambda x : x == row['ISIN']].index
                if len(matching_indices) > 0:
                    matching_index = matching_indices[0]
                    product_long_name = dataframes[2].loc[dataframes[2].index[matching_index], 'Product Long Name']
                else:
                    product_long_name = None
                investment_account_long_name = product_long_name

                attribution_gross = row['Gross Contribution to Index']
                attribution_net = row['Net Contribution to Index']
                opening_allocation = row['Beginning Weight %']
                closing_allocation = row['End Weight %']
                opening_equity_helper_indices = dataframes[1]['Index Name'].loc[lambda x : x=="HFRI-I Liquid Alt UCITS Index\n(Net)"].index
                for index in opening_equity_helper_indices:
                    if reference_day == dataframes[1].loc[dataframes[1].index[index], 'Date']:
                        nav_index = index
                        break
                    else:
                        nav_index = None
                if nav_index != None:
                    opening_equity = dataframes[1].loc[dataframes[1].index[nav_index], 'Previous Day NAV']
                    closing_equity = dataframes[1].loc[dataframes[1].index[nav_index], 'NAV']
                else:
                    opening_equity = None
                    closing_equity = None
                investment_performance = row['% Price Change']
                investment_adj_opening_balance = float(opening_allocation) * float(opening_equity)
                investment_closing_balance = float(closing_allocation) * float(closing_equity)
                portfolio_opening_balance = investment_adj_opening_balance
                portfolio_closing_balance = investment_closing_balance
              
                new_row.append(reference_day)  #Reference Day
                new_row.append(periodicty) #Periodicity
                new_row.append(investor_account_uid) #Investor Account UID
                new_row.append(investor_account_long_name) #Investor Account Long Name
                new_row.append(investment_account_uid) #Investment Account UID
                new_row.append(investment_account_long_name) #Investment Account Long Name
                new_row.append(attribution_gross) #Attribution Gross
                new_row.append(attribution_net) #Attribution Net
                new_row.append(opening_allocation)  #Opening Allocation
                new_row.append(closing_allocation)  #Closing Allocation
                new_row.append(opening_equity) #Opening Equity
                new_row.append(closing_equity) #Closing Equity
                new_row.append(investment_performance) #Investment Performance
                new_row.append(investment_adj_opening_balance) #Investment Adj Opening Balance
                new_row.append(investment_closing_balance) #Investment Closing Balance
                new_row.append(portfolio_opening_balance) #Portfolio Opening Balance
                new_row.append(portfolio_closing_balance) #Portfolio Closing Balance
                data.append(new_row)

        new_dataframe = pd.DataFrame(data, columns=[
            'Reference Day',
            'Periodicity', 
            'Investor Account UID', 
            'Investor Account Long Name',
            'Investment Account UID',
            'Investment Account Long Name',
            'Attribution Gross',
            'Attribution Net',
            'Opening Allocation',
            'Closing Allocation',
            'Opening Equity',
            'Closing Equity',
            'Investment Performance',
            'Investment Adj Opening Balance',
            'Investment Closing Balance',
            'Portfolio Opening Balance',
            'Portfolio Closing Balance'
            ])
        new_dataframe.to_excel(new_filename, sheet_name = new_filename.strip('.xlsx'), index=False)
        return new_dataframe

    def generate_portfolio_valuation_file2(self, old_filename, new_filename):
        reference_date = self.extract_timestamp_from_file(old_filename)
        dataframes = []
        wb = load_workbook(old_filename)
        xls = pd.ExcelFile(old_filename)

        for worksheet in wb.sheetnames:
            dataframes.append(pd.read_excel(xls, worksheet))
        dataframes[1].columns = dataframes[1].columns.str.strip()
        data = []

        #dataframes[0] = constituents
        #dataframes[1] = index data
        for index, row in dataframes[1].iterrows():
            new_row = []
            if reference_date.replace('/', '-') in str(row['Date']):
                if row['ISIN'] == "HFRIILAU":
                    portfolio_account_uid = row['ISIN']
                    account_long_name = row['Index Name']
                    date = row['Date']

                    new_row.append(portfolio_account_uid) #Portfolio Account UID
                    new_row.append(account_long_name) #Account Long Name
                    new_row.append(date) #Date
                    if "HFRI-I Liquid Alt UCITS Index" in row['Index Name']:
                        nav_share = row['NAV']
                        new_row.append(nav_share) #NAV/Share
                    else:
                        new_row.append(None)

                    final = 'True'

                    new_row.append(final) #Final
            if len(new_row) != 0:
                data.append(new_row)

        new_dataframe = pd.DataFrame(data, columns=['Portfolio Account UID', 'Account Long Name', 'Date', 'NAV/Share', 'Final'])
        new_dataframe.to_excel(new_filename, sheet_name=new_filename.strip('.xlsx'), index=False)
        return new_dataframe

    def extract_timestamp_from_file(self, filename):
        timestamp = filename.strip('.xlsx').split('_')[-1]
        month, day, year = timestamp.split('.')
        timestamp = f'{year}/{month}/{day}'
        return timestamp


    def convert_xlsx_to_xml(self, filenames):
         for filename in filenames:
            wb = load_workbook(filename)
            xls = pd.ExcelFile(filename)
            dataframes = []
            for worksheet in wb.sheetnames:
                dataframes.append((pd.read_excel(xls, worksheet), worksheet))
            
            for df, name in dataframes:
                new_file = open(f'{name}.xml', 'w')
                df.columns = df.columns.str.replace(' ', '')
                name = name.strip().replace(' ', '')
                xml_doc = f'<{name}>\n'
                for row in df.iterrows():
                    new_row = '\t<row>\n'
                    for i in range(len(df.columns)):
                        row_data = str(row[1][df.columns[i]]).replace("\n", " ")
                        new_row += f'\t\t<{df.columns[i]}>{row_data}</{df.columns[i]}>\n'
                    new_row+='\t</row>\n'
                    xml_doc+=new_row
                xml_doc += f'</{name}>'
                new_file.write(xml_doc)
                new_file.close()

    def run_query(self, query):
        queryset = self.con.execute(f"SELECT * FROM constituents")
        print(queryset.fetchall()[0])